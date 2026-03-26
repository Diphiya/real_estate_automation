"""
Cirrus Real Estate — Excel Template Populator (FINAL)
Exact mapping to template + KPIs + robust fallbacks
"""

import argparse
import json
import os
import shutil
import sys
from datetime import date

from openpyxl import load_workbook
from openpyxl.styles import Font

BLUE  = Font(color="0000FF")
GREEN = Font(color="008000")

TEMPLATE_DEFAULT = "Case_Study__Aufteiler_.xlsx"


# ── Helper ───────────────────────────────────────────
def write(ws, row, col, value, font=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font


def safe(data, *keys, default=0):
    """Try multiple key names (important for LLM variations)"""
    for k in keys:
        if k in data and data[k] not in [None, ""]:
            return data[k]
    return default


def compute_rent(data):
    """Aggregate rent from tenants if not directly available"""
    if data.get("total_rent"):
        return data["total_rent"]

    tenants = data.get("tenants", [])
    total = 0

    for t in tenants:
        rent = (
            t.get("cold_rent")
            or t.get("rent")
            or t.get("monthly_rent")
            or 0
        )
        try:
            total += float(rent)
        except:
            pass

    return round(total, 2)


# ── Main Populate ────────────────────────────────────
def populate(data: dict, price_sqm: float, config: dict,
             template_path: str = None,
             output_path: str = None,
             price_source: str = "manual") -> str:

    tpl = template_path or TEMPLATE_DEFAULT
    if not os.path.exists(tpl):
        print(f"  ✗ Template not found: {tpl}")
        sys.exit(1)

    out = output_path or f"Cirrus_BusinessCase_{date.today().isoformat()}.xlsx"
    shutil.copy2(tpl, out)

    wb = load_workbook(out)

    # ── Normalize Data ────────────────────────────────
    address = safe(data, "address")
    district = safe(data, "district", "city")
    property_type = safe(data, "objectType", "property_type")
    year = safe(data, "buildYear", "year_built")
    renovation = safe(data, "lastRenovation", "modernization")
    plot = safe(data, "plotSize", "land_area")
    floors = safe(data, "floors")
    attic = safe(data, "atticDeveloped", default=False)
    units = safe(data, "units")
    living_area = safe(data, "totalLivingArea", "living_area")

    purchase_price = safe(data, "purchasePrice", default=0)

    # ── Rent + KPIs ──────────────────────────────────
    rent_monthly = compute_rent(data)
    annual_rent = rent_monthly * 12 if rent_monthly else 0

    estimated_price = data.get("estimated_price", living_area * price_sqm)
    yield_pct = (annual_rent / estimated_price * 100) if estimated_price else 0

    # ── Sheet 1: Stammdaten ──────────────────────────
    sd = wb["INPUT_Stammdaten"]

    write(sd, 5, 5, date.today().strftime("%d.%m.%Y"))

    write(sd, 10, 3, address, font=BLUE)
    write(sd, 11, 3,
          "https://maps.google.com/?q=" + str(address).replace(" ", "+"),
          font=BLUE)
    write(sd, 12, 3, district, font=BLUE)
    write(sd, 13, 3, property_type, font=BLUE)
    write(sd, 14, 3, year, font=BLUE)
    write(sd, 15, 3, renovation, font=BLUE)
    write(sd, 16, 3, plot, font=BLUE)

    write(sd, 20, 3, floors, font=BLUE)
    write(sd, 21, 3, "Ja" if attic else "Nein", font=BLUE)

    write(sd, 32, 3, units, font=BLUE)
    write(sd, 33, 3, living_area, font=BLUE)

    write(sd, 45, 3, purchase_price, font=BLUE)

    # ── Sheet 2: Verkaufseinschätzung ────────────────
    vk = wb["INPUT_Verkaufseinschätzung Mark"]

    write(vk, 24, 6, living_area, font=BLUE)

    tenants = data.get("tenants", [])

    price_font = GREEN if price_source == "check24" else BLUE

    for i, t in enumerate(tenants[:11]):
        r = 9 + i

        write(vk, r, 3, "WE", font=BLUE)
        write(vk, r, 4, t.get("floor"), font=BLUE)
        write(vk, r, 5, t.get("area"), font=BLUE)

        if price_sqm:
            write(vk, r, 7, price_sqm, font=price_font)

    # ── Price Source Label ───────────────────────────
    source_text = "Check24 Auto" if price_source == "check24" else "Manual Input"
    source_font = GREEN if price_source == "check24" else BLUE
    write(vk, 6, 7, f"Price Source: {source_text}", font=source_font)

    # ── Sheet 3: Finanzierung ────────────────────────
    fin_ws = wb["INPUT_Finanzierung"]
    fin = config.get("financing", {})

    write(fin_ws, 10, 3, fin.get("repaymentRatePA", 0.01), font=BLUE)
    write(fin_ws, 11, 3, fin.get("interestRatePA", 0.04), font=BLUE)

    # ── EXTRA: KPI BLOCK (VERY IMPORTANT) ────────────
    # Add below existing content (safe area)
    write(fin_ws, 20, 3, "Monthly Rent", font=BLUE)
    write(fin_ws, 20, 4, rent_monthly, font=BLUE)

    write(fin_ws, 21, 3, "Annual Rent", font=BLUE)
    write(fin_ws, 21, 4, annual_rent, font=BLUE)

    write(fin_ws, 22, 3, "Price per sqm", font=BLUE)
    write(fin_ws, 22, 4, price_sqm, font=BLUE)

    write(fin_ws, 23, 3, "Estimated Price", font=BLUE)
    write(fin_ws, 23, 4, estimated_price, font=BLUE)

    write(fin_ws, 24, 3, "Gross Yield (%)", font=BLUE)
    write(fin_ws, 24, 4, round(yield_pct, 2), font=BLUE)

    wb.save(out)
    return out


# ── CLI ──────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Populate Cirrus Excel template")
    parser.add_argument("data")
    parser.add_argument("--price-sqm", type=float, default=0)
    parser.add_argument("--template", default=TEMPLATE_DEFAULT)
    parser.add_argument("--output")
    parser.add_argument("--config", default="config.json")
    parser.add_argument("--price-source", default="manual")

    args = parser.parse_args()

    with open(args.data, encoding="utf-8") as f:
        data = json.load(f)

    config = {}
    if os.path.exists(args.config):
        with open(args.config, encoding="utf-8") as f:
            config = json.load(f)

    out = populate(
        data,
        args.price_sqm,
        config,
        template_path=args.template,
        output_path=args.output,
        price_source=args.price_source
    )

    print(f"✓ Excel written: {out}")


if __name__ == "__main__":
    main()